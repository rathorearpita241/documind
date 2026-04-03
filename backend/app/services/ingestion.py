"""
Ingestion pipeline: parse → chunk → embed → store in ChromaDB
Supports: PDF (.pdf), Excel (.xlsx), Email (.eml / .txt)
"""
import uuid
from pathlib import Path
from typing import List, Dict, Any

import fitz  # PyMuPDF
import openpyxl
from app.core.config import settings
from app.core.vectorstore import get_collection, get_embedder


def _chunk_text(text: str, size: int = settings.chunk_size, overlap: int = settings.chunk_overlap) -> List[str]:
    words = text.split()
    chunks = []
    i = 0
    while i < len(words):
        chunk = " ".join(words[i:i + size])
        if chunk.strip():
            chunks.append(chunk.strip())
        i += size - overlap
    return chunks


def parse_pdf(path: str) -> List[Dict[str, Any]]:
    doc = fitz.open(path)
    chunks = []
    for page_num, page in enumerate(doc):
        text = page.get_text()
        for chunk in _chunk_text(text):
            chunks.append({"text": chunk, "section": f"Page {page_num + 1}"})
    return chunks


def parse_xlsx(path: str) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    chunks = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h) if h is not None else f"Col{i}" for i, h in enumerate(rows[0])]
        for row_idx, row in enumerate(rows[1:], start=2):
            row_text = " | ".join(
                f"{headers[i]}: {v}"
                for i, v in enumerate(row)
                if v is not None and i < len(headers)
            )
            if row_text.strip():
                chunks.append({"text": row_text, "section": f"Sheet '{sheet_name}' Row {row_idx}"})
    return chunks


def parse_email(path: str) -> List[Dict[str, Any]]:
    with open(path, "r", errors="ignore") as f:
        content = f.read()
    chunks = []
    for chunk in _chunk_text(content):
        chunks.append({"text": chunk, "section": "Email body"})
    return chunks


def ingest_file(file_path: str, doc_name: str, source_type: str, doc_date: str) -> int:
    """Parse, embed, and store a document. Returns chunk count."""
    path = Path(file_path)
    if source_type == "pdf":
        raw_chunks = parse_pdf(str(path))
    elif source_type == "xlsx":
        raw_chunks = parse_xlsx(str(path))
    else:
        raw_chunks = parse_email(str(path))

    if not raw_chunks:
        return 0

    collection = get_collection()
    embedder = get_embedder()

    texts = [c["text"] for c in raw_chunks]
    embeddings = embedder.encode(texts).tolist()

    doc_id = str(uuid.uuid4())
    ids = [f"{doc_id}_{i}" for i in range(len(texts))]
    metadatas = [
        {
            "doc_id": doc_id,
            "doc_name": doc_name,
            "source_type": source_type,
            "section": c["section"],
            "doc_date": doc_date,
        }
        for c in raw_chunks
    ]

    collection.add(ids=ids, embeddings=embeddings, documents=texts, metadatas=metadatas)
    return len(texts)


def list_documents() -> List[Dict]:
    collection = get_collection()
    results = collection.get(include=["metadatas"])
    seen = {}
    for meta in results["metadatas"]:
        did = meta["doc_id"]
        if did not in seen:
            seen[did] = {
                "id": did,
                "name": meta["doc_name"],
                "source_type": meta["source_type"],
                "date_uploaded": meta["doc_date"],
                "chunk_count": 0,
            }
        seen[did]["chunk_count"] += 1
    return list(seen.values())
